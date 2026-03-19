import os
import re
import datetime
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import io

def aggregate_results():
    # クロスプラットフォーム互換性のために、スクリプトの場所に基づいた相対パスを使用します
    input_dir = Path(__file__).parent / "runtime_output"
    if not input_dir.exists():
        print(f"Directory not found: {input_dir}")
        return

    subdirs = sorted([d for d in input_dir.iterdir() if d.is_dir()])
    
    raw_data = []
    print(f"Processing {len(subdirs)} result folders...")
    
    for d in subdirs:
        fields_file = d / "fields.txt"
        img_file = d / "final_ocr.jpg"
        
        row = {
            "No.": len(raw_data) + 1,
            "Registration ID": d.name, 
            "Full Name": "", 
            "Email": "", 
            "Company": "",
            "Image Path": str(img_file) if img_file.exists() else ""
        }
        
        if fields_file.exists():
            try:
                with open(fields_file, "r", encoding="utf-8") as f:
                    content = f.read()
                    
                # key: value をマッチング
                for key, col in [("full_name", "Full Name"), ("email", "Email"), ("company", "Company")]:
                    match = re.search(f"{key}:\\s*(.*)", content)
                    val = match.group(1).strip() if match else ""
                    if val.lower() == "none": val = ""
                    row[col] = val
            except Exception as e:
                print(f"Error reading {fields_file}: {e}")
                    
        raw_data.append(row)

    df = pd.DataFrame(raw_data)
    
    # 統計を計算
    total_count = len(df)
    stats = []
    for col in ["Full Name", "Email", "Company"]:
        missing = df[col].apply(lambda x: not str(x).strip()).sum()
        stats.append({
            "Field": col,
            "Missing": missing,
            "Total": total_count,
            "Missing %": (missing / total_count * 100) if total_count > 0 else 0
        })
    df_stats = pd.DataFrame(stats)

    output_file = input_dir / "report.xlsx"
    
    # 詳細列: No.、Registration ID、Full Name、Email、Company、OCR Image
    cols_to_show = ["No.", "Registration ID", "Full Name", "Email", "Company", "OCR Image"]
    df_details = df.copy()
    df_details["OCR Image"] = "" # プレースホルダー
    df_details = df_details[["No.", "Registration ID", "Full Name", "Email", "Company", "OCR Image"]]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Summary シートを書き込む
        df_stats.to_excel(writer, sheet_name='Summary', index=False)
        summary_sheet = writer.sheets['Summary']
        
        # Summary ヘッダーをフォーマット
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
        # Details シートを書き込む
        df_details.to_excel(writer, sheet_name='Details', index=False)
        details_sheet = writer.sheets['Details']
        
        # Details のフォーマット
        yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        # ヘッダーをフォーマット
        for cell in details_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 列のフォーマットと欠落セルのハイライト
        # df の列: [No., Registration ID, Full Name, Email, Company, OCR Image]
        for row_idx, row_values in enumerate(df.values, start=2):
            # No. (col 1)
            details_sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
            # ID (col 2)
            details_sheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
            
            # フルネーム (インデックス 2)、Email (インデックス 3)、会社名 (インデックス 4) をチェック
            # Full Name は df.values においてインデックス 2 です (No, ID, FN, EM, CO, ImgPath)
            for i in [2, 3, 4]:
                val = row_values[i]
                excel_col = i + 1
                cell = details_sheet.cell(row=row_idx, column=excel_col)
                if not str(val).strip() or str(val).lower() == "nan":
                    cell.fill = yellow_fill
                    cell.value = "MISSING"

            # 画像を挿入 (列 6)
            img_path = row_values[5] # 画像パスはインデックス 5
            if img_path and os.path.exists(img_path):
                try:
                    # サムネイル用に画像をオープンしてリサイズ
                    with PILImage.open(img_path) as pil_img:
                        # 最大高さ 100px
                        pil_img.thumbnail((400, 100))
                        img_byte_arr = io.BytesIO()
                        pil_img.save(img_byte_arr, format='JPEG')
                        img_byte_arr.seek(0)
                        
                        xl_img = XLImage(img_byte_arr)
                        # 画像に基づいて行の高さを計算する必要がある
                        # Excel のデフォルトの行の高さは約 15
                        details_sheet.row_dimensions[row_idx].height = 80 
                        
                        cell_address = f"{get_column_letter(6)}{row_idx}"
                        details_sheet.add_image(xl_img, cell_address)
                except Exception as e:
                    print(f"Error embedding image {img_path}: {e}")
                    details_sheet.cell(row=row_idx, column=6).value = "Img Error"
            else:
                details_sheet.cell(row=row_idx, column=6).value = "No Image"

        # 列幅の自動調整
        for sheet in [summary_sheet, details_sheet]:
            for column in sheet.columns:
                column_letter = get_column_letter(column[0].col_idx)
                if sheet == details_sheet and column_letter == 'F':
                    sheet.column_dimensions[column_letter].width = 50 # OCR Image 列
                    continue
                    
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except: pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    print(f"Aggregation complete! Report saved to: {output_file}")

if __name__ == "__main__":
    aggregate_results()
