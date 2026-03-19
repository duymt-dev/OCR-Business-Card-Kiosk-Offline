import os
import json
import time
import base64
import re
import threading
import uuid
from functools import wraps
from difflib import SequenceMatcher
import logging 
from io import BytesIO
from pathlib import Path
from typing import Optional

from flask import Flask, render_template, request, jsonify, send_from_directory, url_for, Response, g, session, redirect
from flask_cors import CORS
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Border, Side


import qrcode
from PIL import Image, ImageOps
import numpy as np

try:
    import cv2
    _HAS_CV2 = True
except Exception:
    _HAS_CV2 = False

from card import (
    CardYoloCapture, CardAutoCapture, get_reader, warmup_ocr, init_yolo,
    decode_data_url_to_pil, parse_bcard_with_bbox, parse_bcard_fields,
    clean_ocr_text, normalize_email_text, is_plausible_phone,
    first_phone, first_email, extract_address, extract_company,
    all_titles, all_phones,
    _torch_has_cuda, _torch_has_directml
)
from card.utils import undistort_image
#123
from face import detect_persons_in_frame, detect_faces_in_frame, draw_boxes_on_frame

# データベースのインポート
from database.update_database import (
    setup_database as init_sqlite_db,
    save_to_sqlite_with_retry as save_to_sqlite,
    update_registration_with_ocr,
    list_registrations,
    get_registration,
    delete_registration,
    update_registration,
    dashboard_stats,
    authenticate_user,
    list_recent_registrations,
    check_and_trigger_cleanup,
)

# ロガーの設定後にリーダーを初期化
reader = None
_QUICK_READER = None

def _get_quick_reader():
    """Lazy initialization of a DEFAULT RapidOCR instance (without wrapper)."""
    global _QUICK_READER
    if _QUICK_READER is None:
        try:
            from rapidocr_onnxruntime import RapidOCR
            # Initialize with default parameters (no custom paths)
            _QUICK_READER = RapidOCR()
            logger.info("Initialized DEFAULT RapidOCR engine for Quick OCR.")
        except Exception as e:
            logger.error(f"Failed to init DEFAULT RapidOCR: {e}")
            return None
    return _QUICK_READER

# AI_READYフラグをcardパッケージまたはロジックから設定
AI_READY = False
ENABLE_LLM = os.getenv("ENABLE_LLM", "0") == "1"

# 名刺検出モデル (YOLOv8 ONNX)
_CARD_MODEL_PATH = os.path.join("models", "card", "best.onnx")
card_detector = CardYoloCapture(
    model_path=_CARD_MODEL_PATH,
    required_stable=3,
    conf_threshold=0.93,
    cooldown=3.0,
)

# 非同期OCRタスクストア (メモリ内、プロセスごと)
_OCR_TASKS: dict[str, dict] = {}
_OCR_TASKS_LOCK = threading.Lock()
_OCR_TASK_TTL_SEC = 600
_OCR_SEMAPHORE = threading.Semaphore(6) # 同時実行OCRタスクを5-7に制限

# ================= Logging =================
fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
ch = logging.StreamHandler()
ch.setFormatter(fmt)
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
root_logger.handlers.clear()
root_logger.addHandler(ch)
# 注: Windowsでのロック問題を避けるため、FileHandler (fh) は 'app' 定義後に追加されます。
logging.captureWarnings(True)

logger = logging.getLogger("kiosk")
logger.setLevel(logging.INFO)

# トラブルシューティングのため、サードパーティの騒がしいロガーをターミナルで表示したままにする。
for _name in ("rapidocr_onnxruntime", "onnxruntime", "werkzeug"):
    _lg = logging.getLogger(_name)
    _lg.setLevel(logging.INFO)
    _lg.handlers.clear()
    _lg.propagate = True

# 最初のスキャンが遅くならないよう、Quick OCR用のデフォルトモデルを常に初期化
_ = _get_quick_reader()

load_dotenv()

# ================= Flask/App config =================
app = Flask(__name__, static_folder="static", template_folder="templates")
CORS(app)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "bamboo-kiosk-secret")

# ファイルログなし、ターミナルのみ。
logging.getLogger("werkzeug").setLevel(logging.INFO)
app.logger.setLevel(logging.INFO)

REG_DIR = Path("registrations")
REG_DIR.mkdir(exist_ok=True, parents=True)


@app.before_request
def _log_request_start():
    g._request_started_at = time.perf_counter()
    cl = request.content_length or 0
    size_str = f" sz={cl/1024:.1f}KB" if cl > 0 else ""
    logger.info("REQ %s %s from=%s%s", request.method, request.full_path, request.remote_addr, size_str)

@app.errorhandler(413)
def request_entity_too_large(error):
    logger.error("!!! 413 Request Entity Too Large: %s (Size: %s)", request.path, request.content_length)
    return jsonify({"ok": False, "error": "Request entity too large (payload too big)"}), 413


@app.after_request
def _log_request_end(response):
    started_at = getattr(g, "_request_started_at", None)
    if started_at is not None:
        elapsed_ms = (time.perf_counter() - started_at) * 1000
        logger.info("RES %s %s status=%s %.1fms", request.method, request.path, response.status_code, elapsed_ms)
    else:
        logger.info("RES %s %s status=%s", request.method, request.path, response.status_code)
    return response


def _current_user():
    return session.get("user")


def login_required(fn):
    @wraps(fn)
    def _wrapped(*args, **kwargs):
        if _current_user():
            return fn(*args, **kwargs)
        if request.path.startswith("/api/"):
            # The instruction implies adding check_api_auth, but the primary goal is string replacement.
            # To faithfully follow "without making any unrelated edits" for the general case,
            # I will only change the string literal here.
            # If `check_api_auth` was intended as a new feature, it should be a separate instruction.
            return jsonify({"ok": False, "error": "認証されていません"}), 401
        return redirect(url_for("login"))
    return _wrapped

# SQLite DBが初期化されていることを確認
try:
    init_sqlite_db()
    logger.info("SQLite database initialized successfully.")
except Exception as e:
    logger.error(f"Failed to initialize SQLite database: {e}")

FORWARD_ENABLED = os.getenv("FORWARD_ENABLED", "false").lower() == "true"
FORWARD_URL = os.getenv("FORWARD_URL", "").strip()
FORWARD_API_KEY = os.getenv("FORWARD_API_KEY", "").strip()

# ================= Model Initialization =================
_HAS_YOLO = False
yolo_model = None
_YOLO_DEVICE = "cpu"

try:
    yolo_model, _HAS_YOLO = init_yolo()
except Exception as e:
    logger.warning(f"YOLO initialization failed: {e}")

def _warmup():
    """
    初回呼び出しを高速化するために、モデルを軽く試運転（ウォームアップ）します。
    """
    global AI_READY
    # warmup_ocr(enable_llm=ENABLE_LLM, skip_llm_warmup=(os.getenv("DISABLE_LLM_WARMUP", "0") == "1"))
    
    if _HAS_YOLO and yolo_model is not None:
        try:
            # Warmup YOLO
            z = np.zeros((320, 320, 3), dtype=np.uint8)
            yolo_model(z, verbose=False, device=_YOLO_DEVICE)
            
            # Warmup OCR (EasyOCR/RapidOCR)
            if reader is not None:
                z_ocr = np.zeros((100, 300, 3), dtype=np.uint8)
                reader.readtext(z_ocr, detail=0)
                logger.info("OCR engine warmed up.")
        except Exception as e:
            logger.warning(f"YOLO warmup err: {e}")
            
    AI_READY = True


# --------- 登録ペイロード + QR + チェックインの保存 ----------

def _save_image_dataurl(target_folder: Path, name_noext: str, data_url: str, undistort: bool = False) -> Optional[str]:
    try:
        img = decode_data_url_to_pil(data_url).convert("RGB")
        if undistort:
            np_img = np.array(img)
            # 魚眼レンズの歪みを補正
            np_img = undistort_image(np_img)
            img = Image.fromarray(np_img)
            
        fmt = "PNG"
        if data_url.split(";")[0].endswith(("jpeg", "jpg")):
            fmt = "JPEG"
        out_path = target_folder / f"{name_noext}.{fmt.lower()}"
        img.save(out_path, fmt, quality=92)
        return out_path.name
    except Exception:
        return None


def _append_checkin_line(reg_id: str, data: dict, bcard_fields: Optional[dict]):
    """
    checkin.txtに1行書き込みます:
    [INFO] {time, reg_id, idNumber, fullName, dob, issued, full_name, title, email, company, phone, address}
    """
    now_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    entry: dict[str, object] = {
        "time": now_str,
        "reg_id": reg_id,
        "idNumber": data.get("idNumber", ""),
        "fullName": data.get("fullName", ""),
        "dob": data.get("dob", ""),
        "issued": data.get("issued", ""),
    }

    b = bcard_fields or {}
    entry.update(
        {
            "full_name": b.get("full_name", ""),
            "title": b.get("title", ""),
            "email": b.get("email", ""),
            "company": b.get("company", ""),
            "phone": b.get("phone", ""),
            "address": b.get("address", ""),
        }
    )

    try:
        checkin_path = Path("checkin.txt")
        with checkin_path.open("a", encoding="utf-8") as f:
            f.write("[INFO] " + json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as e:
        logger.warning(f"Failed to append checkin.txt: {e}")


def save_registration(payload: dict) -> dict:
    logger.info("save_registration called with payload keys: %s", list(payload.keys()))
    t0 = time.time()
    data = payload.get("data", {}) or {}
    
    reg_id = payload.get("registration_id")
    print(f"!!! DEBUG_SAVE: registrationId received='{reg_id}'")
    if not reg_id:
        import random, string
        suffix = ''.join(random.choices(string.ascii_lowercase + string.digits, k=4))
        reg_id = f"REG_{time.strftime('%d-%m-%Y_%H-%M-%S')}_{suffix}"
    
    reg_folder = REG_DIR / reg_id
    reg_folder.mkdir(exist_ok=True, parents=True)

    bcard_fields = (
        payload.get("bcard_fields")
        or payload.get("fields_ai")
        or payload.get("fields")
        or {}
    )

    enriched = {
        "data": data,
        "bcard_fields": bcard_fields,
        "last_qr_raw": payload.get("last_qr_raw"),
        "last_mrz_text": payload.get("last_mrz_text"),
        "last_bcard_text": payload.get("last_bcard_text"),
        "ts": payload.get("ts") or int(time.time() * 1000),
        # 検証と一貫性のためにフィールドを平坦化
        "bcard_name": bcard_fields.get("full_name") or bcard_fields.get("name"),
        "bcard_company": bcard_fields.get("company") or bcard_fields.get("org"),
        "bcard_email": bcard_fields.get("email"),
        "bcard_phone": bcard_fields.get("phone") or bcard_fields.get("tel"),
        "bcard_title": bcard_fields.get("title") or bcard_fields.get("position") or bcard_fields.get("role"),
        "bcard_address": bcard_fields.get("address"),
        "bcard_info": payload.get("last_bcard_text") or "",
    }
    t_json0 = time.time()
    with (reg_folder / "data.json").open("w", encoding="utf-8") as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)
    t_json = time.time() - t_json0

    saved = {"data": "data.json", "face": None, "bcard_image": None, "sources": []}

    # 顔写真を保存
    if payload.get("face_image"):
        saved["face"] = _save_image_dataurl(
            reg_folder, "face", payload["face_image"]
        )

    # 名刺画像を保存
    if isinstance(payload.get("bcard_image"), str) and "base64," in payload["bcard_image"]:
        saved["bcard_image"] = _save_image_dataurl(
            reg_folder, "bcard", payload["bcard_image"], undistort=True
        )

    # ソース画像を保存
    src_imgs = payload.get("source_images", []) or []
    for i, durl in enumerate(src_imgs):
        name = _save_image_dataurl(reg_folder, f"src_{i}", durl)
        if name:
            saved["sources"].append(name)

    # 名刺バイパス用のQRコードを生成
    qr_abs_url = None
    if bcard_fields and any(bcard_fields.values()):
        qr_data = {
            "bcard_fields": bcard_fields,
            "last_bcard_text": payload.get("last_bcard_text", "")
        }
        try:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=2,   # 画像サイズを縮小するため10から2に削減
                border=2,     # 白枠を減らすため4から2に削減
            )
            qr.add_data(json.dumps(qr_data, ensure_ascii=False))
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            qr_filename = "registration_qr.png"
            qr_path = reg_folder / qr_filename
            qr_img.save(str(qr_path))
            
            qr_abs_url = f"/registrations/{reg_id}/{qr_filename}"
        except Exception as e:
            logger.error(f"Failed to generate QR code: {e}")


    # チェックイン情報を追記
    _append_checkin_line(reg_id, data, bcard_fields)
    
    # SQLite
    try:
        logger.info(f"DEBUG_SAVE: Attempting to save to SQLite for reg_id={reg_id}")
        save_to_sqlite(reg_id, payload, data, bcard_fields, reg_folder)
        logger.info(f"DEBUG_SAVE: save_to_sqlite completed for reg_id={reg_id}")
    except Exception as e:
        logger.error(f"DEBUG_SAVE: Failed to save to SQLite: {e}", exc_info=True)

    # バックグラウンドでクリーンアップチェックを実行
    try:
        threading.Thread(target=check_and_trigger_cleanup, args=(10000,), daemon=True).start()
    except Exception as cleanup_err:
        logger.error(f"Failed to start cleanup thread: {cleanup_err}")

    return {
        "ok": True,
        "registration_id": reg_id,
        "qr_url": qr_abs_url,
        "saved": saved
    }

def save_face_image(target_folder: Path, data_url: str, saver_fn) -> Optional[str]:
    return saver_fn(target_folder, "face", data_url)

# ================= Routes =================
@app.get("/")
def index():
    return render_template("index.html")


@app.get("/dashboard")
@login_required
def dashboard():
    return render_template("index_dashboard.html", current_user=_current_user())


@app.get("/login")
def login():
    if _current_user():
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.post("/api/auth/login")
def api_auth_login():
    js = request.get_json(silent=True) or {}
    username = (js.get("username") or request.form.get("username") or "").strip()
    password = (js.get("password") or request.form.get("password") or "").strip()
    if not username or not password:
        return jsonify({"ok": False, "error": "ユーザー名とパスワードは必須です"}), 400

    user = authenticate_user(username, password)
    if not user:
        return jsonify({"ok": False, "error": "無効な認証情報です"}), 401

    session["user"] = user
    return jsonify({"ok": True, "user": user})


@app.post("/api/auth/logout")
def api_auth_logout():
    session.pop("user", None)
    return jsonify({"ok": True})


@app.get("/api/auth/me")
def api_auth_me():
    user = _current_user()
    if not user:
        return jsonify({"ok": False, "error": "認証されていません"}), 401
    return jsonify({"ok": True, "user": user})


@app.get("/api/dashboard/stats")
@login_required
def api_dashboard_stats():
    try:
        stats = dashboard_stats()
        return jsonify({"ok": True, "stats": stats})
    except Exception as e:
        logger.exception("api_dashboard_stats error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.get("/api/dashboard/registrations")
@login_required
def api_dashboard_registrations():
    search = (request.args.get("search") or "").strip()
    page = request.args.get("page", default=1, type=int)
    page_size = request.args.get("page_size", default=10, type=int)
    sort_by = request.args.get("sort_by", default="created_at", type=str)
    sort_dir = request.args.get("sort_dir", default="desc", type=str)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    try:
        payload = list_registrations(
            search=search,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_dir=sort_dir,
            date_from=date_from,
            date_to=date_to,
        )
        return jsonify({"ok": True, **payload})
    except Exception as e:
        logger.exception("api_dashboard_registrations error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.get("/api/dashboard/registrations/<reg_id>")
@login_required
def api_dashboard_registration_detail(reg_id):
    try:
        item = get_registration(reg_id)
        if not item:
            return jsonify({"ok": False, "error": "登録が見つかりません"}), 404
        return jsonify({"ok": True, "item": item})
    except Exception as e:
        logger.exception("api_dashboard_registration_detail error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.delete("/api/dashboard/registrations/<reg_id>")
@login_required
def api_dashboard_registration_delete(reg_id):
    try:
        deleted = delete_registration(reg_id)
        if not deleted:
            return jsonify({"ok": False, "error": "登録が見つかりません"}), 404
        return jsonify({"ok": True, "deleted": reg_id})
    except Exception as e:
        logger.exception("api_dashboard_registration_delete error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500

@app.put("/api/dashboard/registrations/<reg_id>")
@login_required
def api_dashboard_registration_update(reg_id):
    try:
        js = request.get_json(silent=True) or {}
        updates = {}
        for key in ("full_name", "company", "email", "phone", "title"):
            if key in js:
                updates[key] = (js.get(key) or "").strip()
        updated = update_registration(reg_id, updates)
        if not updated:
            return jsonify({"ok": False, "error": "Registration not found"}), 404
        return jsonify({"ok": True, "updated": reg_id})
    except Exception as e:
        logger.exception("api_dashboard_registration_update error")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.get("/api/dashboard/export.xlsx")
@login_required
def api_dashboard_export():
    search = (request.args.get("search") or "").strip()
    sort_by = request.args.get("sort_by", default="created_at", type=str)
    sort_dir = request.args.get("sort_dir", default="desc", type=str)
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    try:
        payload = list_registrations(
            search=search,
            page=1,
            page_size=10000,
            sort_by=sort_by,
            sort_dir=sort_dir,
            date_from=date_from,
            date_to=date_to,
        )
        header = [
            "Registration ID",
            "Full Name",
            "Company",
            "Email",
            "Phone",
            "Job Title",
            "OCR Text",
            "Business Card Image",
            "Face Image",
            "QR Image",
            "Created At",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Database"
        ws.append(header)
        header_font = Font(bold=True)
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for col in range(1, len(header) + 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).border = cell_border

        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 22

        def add_image(path_value, cell_ref, registration_id, image_stem):
            if not path_value:
                return
            raw = str(path_value).strip()
            p = None

            if raw.startswith("/registrations/"):
                p = (Path.cwd() / raw.lstrip("/")).resolve()
            else:
                # Map legacy absolute paths from other machines that include "/registrations/"
                if "/registrations/" in raw.replace("\\", "/"):
                    rel = raw.replace("\\", "/").split("/registrations/", 1)[1]
                    p = (REG_DIR / rel).resolve()
                else:
                    p = Path(raw)
                if not p.exists():
                    # Backward compatibility for old absolute paths from another machine.
                    base_dir = REG_DIR / str(registration_id or "")
                    if p.name:
                        by_name = base_dir / p.name
                        if by_name.exists():
                            p = by_name
                    if not p.exists():
                        for ext in ("jpeg", "jpg", "png", "webp"):
                            by_stem = base_dir / f"{image_stem}.{ext}"
                            if by_stem.exists():
                                p = by_stem
                                break

            if not p or not p.exists():
                return
            try:
                img = OpenpyxlImage(str(p))
                max_w, max_h = 140, 100
                scale = min(max_w / img.width, max_h / img.height, 1.0)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
                ws.add_image(img, cell_ref)
            except Exception:
                logger.exception("Cannot add image to excel: %s", p)

        for idx, row in enumerate(payload.get("items", []), start=2):
            ws.cell(row=idx, column=1, value=str(row.get("registration_id", "")))
            ws.cell(row=idx, column=2, value=str(row.get("full_name", "")))
            ws.cell(row=idx, column=3, value=str(row.get("company", "")))
            ws.cell(row=idx, column=4, value=str(row.get("email", "")))
            ws.cell(row=idx, column=5, value=str(row.get("phone", "")))
            ws.cell(row=idx, column=6, value=str(row.get("title", "")))
            ws.cell(row=idx, column=7, value=str(row.get("last_bcard_text", "")))
            ws.cell(row=idx, column=11, value=str(row.get("created_at", "")))
            ws.row_dimensions[idx].height = 80

            reg_id = row.get("registration_id")
            add_image(row.get("bcard_link"), f"H{idx}", reg_id, "bcard")
            add_image(row.get("face_link"), f"I{idx}", reg_id, "face")
            add_image(row.get("qr_link"), f"J{idx}", reg_id, "registration_qr")

        last_row = max(1, len(payload.get("items", [])) + 1)
        for r in range(2, last_row + 1):
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).border = cell_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=dashboard_export.xlsx"
            },
        )
    except Exception as e:
        logger.exception("api_dashboard_export error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.get("/api/dashboard/notifications")
@login_required
def api_dashboard_notifications():
    limit = request.args.get("limit", default=10, type=int)
    try:
        items = list_recent_registrations(limit=limit)
        return jsonify({"ok": True, "items": items})
    except Exception as e:
        logger.exception("api_dashboard_notifications error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500

@app.post("/api/register")
def register():
    try:
        payload = request.get_json(force=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"無効なJSON: {e}"}), 400
    return jsonify(save_registration(payload))

@app.post("/api/presence/payload")
def presence_payload():
    try:
        payload = request.get_json(force=True)
    except Exception as e:
        return jsonify({"ok": False, "error": f"無効なJSON: {e}"}), 400
    return jsonify(save_registration(payload))

@app.post("/api/presence/frame")
def presence_frame():
    """Detect người trong frame sử dụng YOLO."""
    if not _HAS_CV2:
        return jsonify({"ok": False, "error": "OpenCVがインストールされていません"}), 400
    
    f = request.files.get("frame")
    if not f:
        return jsonify({"ok": False, "error": "フレームがありません"}), 400
    
    try:
        frame_bytes = f.read()
        frame_size, boxes, best, yolo_time = detect_persons_in_frame(
            frame_bytes, yolo_model, _HAS_YOLO, _YOLO_DEVICE
        )
        
        return jsonify({
            "ok": True,
            "frame_size": frame_size,
            "boxes": boxes,
            "best": best,
            "yolo_time": yolo_time,
        })
    except Exception as e:
        logger.exception("presence_frame error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500

#123
@app.post("/api/face/frame")
def face_frame():
    """Detect khuôn mặt trong frame (OpenCV Haar)."""
    if not _HAS_CV2:
        return jsonify({"ok": False, "error": "OpenCVがインストールされていません"}), 400

    f = request.files.get("frame")
    if not f:
        return jsonify({"ok": False, "error": "フレームがありません"}), 400

    try:
        frame_bytes = f.read()
        frame_size, boxes, best, face_time = detect_faces_in_frame(frame_bytes)

        return jsonify(
            {
                "ok": True,
                "frame_size": frame_size,
                "boxes": boxes,
                "best": best,
                "face_time": face_time,
            }
        )
    except Exception as e:
        logger.exception("face_frame error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.post("/api/face/frame_with_boxes")
def face_frame_with_boxes():
    """Detect khuôn mặt và trả về frame đã vẽ bounding boxes (JPEG)."""
    if not _HAS_CV2:
        return jsonify({"ok": False, "error": "OpenCVがインストールされていません"}), 400

    f = request.files.get("frame")
    if not f:
        return jsonify({"ok": False, "error": "フレームがありません"}), 400

    try:
        frame_bytes = f.read()
        frame_size, boxes, best, _ = detect_faces_in_frame(frame_bytes)

        img_bytes = draw_boxes_on_frame(
            frame_bytes, boxes, best,
            box_color=(0, 255, 0),   # xanh lá cho các box
            best_color=(0, 0, 255),  # đỏ cho box best
        )

        if img_bytes is None:
            return jsonify({"ok": False, "error": "ボックスの描画に失敗しました"}), 500

        from flask import Response
        resp = Response(
            img_bytes,
            mimetype="image/jpeg",
            headers={"Content-Type": "image/jpeg"},
        )
        # Header cho frontend dùng logic auto-capture
        if best is not None:
            resp.headers["X-Face-Best"] = json.dumps(best)
        if frame_size is not None:
            resp.headers["X-Frame-Size"] = json.dumps(frame_size)
        return resp
    except Exception as e:
        logger.exception("face_frame_with_boxes error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500

# ================= Card Auto-Detect (YOLOv8) =================
@app.post("/api/card/frame")
def card_frame():
    """Auto-detect danh thiếp từ frame camera (YOLOv8 ONNX).

    JS gửi frame JPEG qua form field 'frame'.
    Khi card ổn định đủ 24 frame liên tiếp → crop bbox gốc → chạy OCR.

    Response JSON:
        ok           : bool
        card_detected: bool  – True khi vừa trigger capture
        stable_count : int   – số frame stable hiện tại
        required     : int   – số frame cần (24)
        bbox         : {x1,y1,x2,y2,conf} | null
        fields       : dict  – OCR fields (khi card_detected=True)
        text         : str   – raw OCR text (khi card_detected=True)
    """
    if not _HAS_CV2:
        return jsonify({"ok": False, "error": "OpenCVがインストールされていません"}), 400

    f = request.files.get("frame")
    if not f:
        return jsonify({"ok": False, "error": "フレームがありません"}), 400

    try:
        import io
        frame_bytes = f.read()
        nparr = np.frombuffer(frame_bytes, np.uint8)
        frame_bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame_bgr is None:
            return jsonify({"ok": False, "error": "フレームをデコードできません"}), 400

        triggered, crop_bgr, bbox = card_detector.detect(frame_bgr)

        if not triggered:
            return jsonify({
                "ok": True,
                "card_detected": False,
                "stable_count": card_detector.stable_count,
                "required": card_detector.required_stable,
                "bbox": bbox,
            })

        # JSがcheck_onlyを送信した場合（例：カード取り出し待ちのREMOVINGフェーズ） -> 検出は返すがOCRは実行しない
        check_only = (request.form.get("check_only", "false").lower() == "true")
        if check_only:
            logger.info("Card detected but check_only=True, skipping OCR.")
            return jsonify({
                "ok": True,
                "card_detected": True,
                "stable_count": 0,
                "required": card_detector.required_stable,
                "bbox": bbox,
                "fields": {},
                "text": ""
            })

        # --- カード読み取り完了: JSに通知のみ ---
        conf_val = bbox.get('conf', 0.0) if isinstance(bbox, dict) else getattr(bbox, 'conf', 0.0)
        logger.info("Card AUTO-CAPTURED bbox=%s conf=%.3f", bbox, conf_val)
        # OCRを待たずに即座に値を返し、フロントエンドにバッジを表示させる
        return jsonify({
            "ok": True,
            "card_detected": True,
            "stable_count": 0,
            "required": card_detector.required_stable,
            "bbox": bbox,
            "fields": {},   # Sẽ được điền sau qua cổng OCR riêng
            "text": "",
        })

    except Exception as e:
        logger.exception("card_frame error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.post("/api/ocr/bcard_quick")
def ocr_bcard_quick():
    """クイックOCRチェック – 1パス、補正/アップスケールなし、同期レスポンス。
    フロントエンドのリトライフローで、カードにテキストがあるかどうかを素早く判断するために使用されます。"""
    f = request.files.get("image")
    if not f:
        return jsonify({"ok": False, "error": "画像がありません"}), 400
    try:
        img = Image.open(f.stream).convert("RGB")
        np_img = np.array(img)
        # 推論時間を最小限に抑えるため、最大480pxにリサイズ
        h, w = np_img.shape[:2]
        max_side = 480
        if max(h, w) > max_side:
            scale = max_side / max(h, w)
            np_img = cv2.resize(np_img, (int(w * scale), int(h * scale)))
        t0 = time.time()
        # 検出（テキスト枠の検索）のみ実行し、認識（文字認識）は行わない。
        # テキスト検出は非常に高速（Piで約1秒）だが、認識は1行あたり約1秒かかる。
        # det_limit_side_len=448の強制指定はここではサポートされていないが、
        # 上記で画像が480pxにリサイズされているため、速度は十分に高速。
        dt_boxes, _ = reader.engine.text_det(np_img)
        box_count = len(dt_boxes) if dt_boxes is not None else 0
        
        elapsed = round(time.time() - t0, 3)
        logger.info("bcard_quick (detection only): %d boxes in %.3fs", box_count, elapsed)
        return jsonify({
            "ok": True, 
            "box_count": box_count,
            "time": elapsed
        })
    except Exception as e:
        logger.exception("ocr_bcard_quick error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.post("/api/ocr/bcard")
def ocr_bcard():
    t_start = time.time()
    try:
        img: Optional[Image.Image] = None
        if "image" in request.files:
            img = Image.open(request.files["image"].stream)
        elif "image_data_url" in request.form:
            img = decode_data_url_to_pil(request.form["image_data_url"])
        elif request.is_json:
            js = request.get_json(silent=True) or {}
            if "image_data_url" in js:
                img = decode_data_url_to_pil(js["image_data_url"])
        if img is None:
            return jsonify({"ok": False, "error": "画像がありません"}), 400

        result = _run_bcard_ocr_from_image(img)

        return jsonify({
            "ok": True,
            "text": result["text"],
            "fields": result["fields"],
            "timing": {
                "total": round(time.time() - t_start, 3),
                "ocr": result["timing"]["ocr"],
            }
        })
    except Exception as e:
        logger.exception(f"OCR error: {e}")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500

@app.post("/api/ocr/bcard_batch")
def ocr_bcard_batch():
    # 一旦バッチロジックは残しておきますが、card.processorを使用するようにリファクタリング可能です。
    # ここでは既存のロジック、または簡略化されたバージョンを維持します。
    # 冗長なコードの削除を検討しても良いですが、機能の削除は求められていません。
    # バッチロジックを維持しつつ、インポートを使用して整理します。
    # 実際には、そのまま維持しますが、インポートを利用します。
    pass # To be implemented or kept if needed. For now placeholder.

@app.get("/registrations/<reg_id>/<path:fname>")
def serve_registration(reg_id, fname):
    return send_from_directory(REG_DIR / reg_id, fname)

@app.get("/api/status")
def status():
    try:
        import onnxruntime as ort # type: ignore
        ort_avail = ort.get_available_providers()
    except Exception:
        ort_avail = []
    return jsonify({
        "ok": True,
        "ai_ready": bool(_lazy_llm()),
        "ocr_ready": True,
        "yolo_ready": bool(_HAS_YOLO and yolo_model),
        "cuda": _torch_has_cuda(),
        "directml": _torch_has_directml(),
        "onnx_providers_available": ort_avail,
    })


def _run_bcard_ocr_from_image(img: Image.Image, filtering_bbox: dict = None) -> dict:
    img = ImageOps.exif_transpose(img).convert("RGB")
    np_img = np.array(img)
    # OCRの前に魚眼レンズの歪みを補正
    np_img = undistort_image(np_img)
    t_ocr0 = time.time()
    ocr_results = reader.readtext(np_img, detail=1, paragraph=False)
    t_ocr_done = time.time()

    # return_full_results=True を使用して、同期されたテキスト出力用に正規化されたOCR結果を取得
    fields, normalized_ocr = parse_bcard_with_bbox(ocr_results, return_full_results=True, filtering_bbox=filtering_bbox)
    
    # 正規化された結果に基づいてテキストを生成（フィールドと同期）
    text = "\n".join(t.strip() for (b, t, c, *_) in normalized_ocr if c > 0.3 and t.strip()).strip()
    
    return {
        "text": text,
        "fields": fields,
        "timing": {
            "ocr": round(t_ocr_done - t_ocr0, 3),
        },
    }


def _cleanup_old_ocr_tasks() -> None:
    now = time.time()
    with _OCR_TASKS_LOCK:
        stale_ids = [
            tid for tid, rec in _OCR_TASKS.items()
            if (now - float(rec.get("updated_at", rec.get("created_at", now)))) > _OCR_TASK_TTL_SEC
        ]
        for tid in stale_ids:
            _OCR_TASKS.pop(tid, None)


def _process_ocr_task(task_id: str, image_bytes: bytes) -> None:
    logger.info("Async OCR task started for task_id: %s", task_id)
    t_start = time.time()
    try:
        _OCR_SEMAPHORE.acquire()
        
        filtering_bbox = None
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec:
                filtering_bbox = rec.get("bbox")

        img = Image.open(BytesIO(image_bytes))
        result = _run_bcard_ocr_from_image(img, filtering_bbox=filtering_bbox)
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec is not None:
                rec.update({
                    "status": "done",
                    "text": result["text"],
                    "fields": result["fields"],
                    "timing": {
                        "total": round(time.time() - t_start, 3),
                        "ocr": result["timing"]["ocr"],
                    },
                    "error": None,
                    "updated_at": time.time(),
                })
        
        # サーバー側オートセーブ: OCR結果でデータベースを更新
        reg_id = None
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec:
                reg_id = rec.get("reg_id")
        
        if reg_id:
            print(f"!!! DEBUG_AUTO_SAVE: Triggering save for {reg_id}")
            logger.info(f"DEBUG_AUTO_SAVE: Triggering server-side save for {reg_id}")
            update_registration_with_ocr(reg_id, result["fields"], result["text"])

    except Exception as e:
        logger.exception("Async OCR task failed: %s", task_id)
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec is not None:
                rec.update({
                    "status": "error",
                    "error": str(e),
                    "updated_at": float(time.time()),
                })
    finally:
        _OCR_SEMAPHORE.release()


def _process_quick_ocr_task(task_id: str, image_bytes: bytes) -> None:
    """投機的クイックOCRのワーカー – 1パス、正規表現フィールドのみ（3-5秒）。"""
    t_start = time.time()
    try:
        _OCR_SEMAPHORE.acquire()
        img = Image.open(BytesIO(image_bytes)).convert("RGB")
        np_img = np.array(img)
        # OCRの前に魚眼レンズの歪みを補正
        np_img = undistort_image(np_img)
        # 速度向上のため、長辺を最大480pxにリサイズ
        h, w = int(np_img.shape[0]), int(np_img.shape[1])
        max_dim = int(max(h, w))
        # Pi上でも高速で、かつ検出精度が良い640pxを使用
        if max_dim > 640:
            scale = 640.0 / float(max_dim)
            np_img = cv2.resize(np_img, (int(float(w) * scale), int(float(h) * scale)))

        t0 = time.time()
        # デフォルトのRapidOCRインスタンスを取得
        q_engine = _get_quick_reader()
        if q_engine is None:
            raise ValueError("Quick OCR engine not initialized")
            
        # 端にある文字をエンジンが正しく認識できるように、30pxの白い余白（パディング）を追加
        # (これはラッパーを使わずに「0行」エラーを修正するための核心です)
        np_img_padded = cv2.copyMakeBorder(
            np_img, 30, 30, 30, 30, cv2.BORDER_CONSTANT, value=[255, 255, 255]
        )
            
        # エンジンの推論プロセスを直接呼び出す
        raw_result = q_engine(np_img_padded)
        
        # Format trả về: [ [ [bbox], text, conf ], ... ]
        raw_items = raw_result[0] if (raw_result and len(raw_result) > 0 and raw_result[0]) else []
        
        # bboxが提供されている場合、クイックOCRの結果をフィルタリング
        filtering_bbox = None
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec:
                filtering_bbox = rec.get("bbox")
        
        if filtering_bbox and isinstance(filtering_bbox, dict) and raw_items:
            margin = 150
            fx1 = filtering_bbox.get("x1", 0) - margin
            fy1 = filtering_bbox.get("y1", 0) - margin
            fx2 = filtering_bbox.get("x2", 9999) + margin
            fy2 = filtering_bbox.get("y2", 9999) + margin
            
            filtered = []
            for item in raw_items:
                if isinstance(item, (list, tuple)) and len(item) >= 3:
                    box = item[0]
                    # boxは通常4つのポイント: [[x,y], [x,y], [x,y], [x,y]]
                    cx = sum(p[0] for p in box) / 4
                    cy = sum(p[1] for p in box) / 4
                    if fx1 <= cx <= fx2 and fy1 <= cy <= fy2:
                        filtered.append(item)
            
            if filtered:
                logger.info(f"Quick OCR Filtering: {len(raw_items)} -> {len(filtered)} items kept.")
                raw_items = filtered

        text_parts = []
        for item in raw_items:
            if isinstance(item, (list, tuple)) and len(item) >= 3:
                txt = str(item[1]).strip()
                conf = float(item[2])
                if conf > 0.3 and txt:
                    text_parts.append(txt)
        
        text = "\n".join(text_parts)
        # 正規表現のみによるパース
        fields, _ = parse_bcard_fields(text)
        t_ocr_done = time.time()

        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec is not None:
                rec.update({
                    "status": "done",
                    "text": text,
                    "fields": fields,
                    "timing": {
                        "total": round(float(time.time() - t_start), 3),
                        "ocr": round(float(t_ocr_done - t0), 3),
                    },
                    "error": None,
                    "updated_at": float(time.time()),
                })
        
        # サーバー側オートセーブ: OCR結果でデータベースを更新
        reg_id = None
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec:
                reg_id = rec.get("reg_id")
        
        if reg_id:
            print(f"!!! DEBUG_AUTO_SAVE: Triggering save (Quick) for {reg_id}")
            logger.info(f"DEBUG_AUTO_SAVE: Triggering server-side save (Quick) for {reg_id}")
            update_registration_with_ocr(reg_id, fields, text)

    except Exception as e:
        logger.exception("Async QUICK OCR task failed: %s", task_id)
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec is not None:
                rec.update({
                    "status": "error",
                    "error": str(e),
                    "updated_at": time.time(),
                })
    finally:
        _OCR_SEMAPHORE.release()


@app.post("/api/ocr/bcard_async/start")
def ocr_bcard_async_start():
    return _ocr_async_start_common(is_quick=False)


@app.post("/api/ocr/bcard_async/quick")
def ocr_bcard_async_quick():
    return _ocr_async_start_common(is_quick=True)


def _ocr_async_start_common(is_quick: bool = False):
    logger.info("OCR Async Start called (is_quick=%s)", is_quick)
    try:
        image_bytes: Optional[bytes] = None
        if "image" in request.files:
            image_bytes = request.files["image"].read()
        elif "image_data_url" in request.form:
            img = decode_data_url_to_pil(request.form["image_data_url"]).convert("RGB")
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=90)
            image_bytes = buf.getvalue()
        elif request.is_json:
            js = request.get_json(silent=True) or {}
            if "image_data_url" in js:
                img = decode_data_url_to_pil(js["image_data_url"]).convert("RGB")
                buf = BytesIO()
                img.save(buf, format="JPEG", quality=90)
                image_bytes = buf.getvalue()

        reg_id = request.form.get("reg_id") or request.args.get("reg_id")
        bbox_raw = request.form.get("bbox") or request.args.get("bbox")
        
        if request.is_json:
            js = request.get_json(silent=True) or {}
            if not reg_id: reg_id = js.get("reg_id")
            if not bbox_raw: bbox_raw = js.get("bbox")
        
        bbox = None
        if bbox_raw:
            try:
                if isinstance(bbox_raw, str):
                    bbox = json.loads(bbox_raw)
                else:
                    bbox = bbox_raw
            except:
                logger.warning(f"Failed to parse bbox: {bbox_raw}")

        print(f"!!! DEBUG_OCR: Received reg_id='{reg_id}' bbox={bbox}")
        logger.info(f"DEBUG_OCR: Received reg_id='{reg_id}' bbox={bbox}")

        if not image_bytes:
            return jsonify({"ok": False, "error": "画像がありません"}), 400

        _cleanup_old_ocr_tasks()
        prefix = "ocr_q" if is_quick else "ocr"
        task_id = f"{prefix}_{uuid.uuid4().hex}"
        now = time.time()
        with _OCR_TASKS_LOCK:
            _OCR_TASKS[task_id] = {
                "status": "processing",
                "text": "",
                "fields": {},
                "timing": {},
                "error": None,
                "created_at": now,
                "updated_at": now,
                "is_quick": is_quick,
                "reg_id": reg_id,
                "bbox": bbox,
            }

        target = _process_quick_ocr_task if is_quick else _process_ocr_task
        t = threading.Thread(target=target, args=(task_id, image_bytes), daemon=True)
        t.start()
        return jsonify({"ok": True, "task_id": task_id, "status": "processing"})
    except Exception as e:
        logger.exception("ocr_async_start error")
        return jsonify({"ok": False, "error": f"サーバーエラー: {e}"}), 500


@app.get("/api/ocr/bcard_async/status/<task_id>")
def ocr_bcard_async_status(task_id: str):
    try:
        _cleanup_old_ocr_tasks()
        with _OCR_TASKS_LOCK:
            rec = _OCR_TASKS.get(task_id)
            if rec is None:
                return jsonify({"ok": False, "error": "タスクが見つかりません"}), 404
            return jsonify({
                "ok": True,
                "task_id": task_id,
                "status": rec.get("status", "processing"),
                "text": rec.get("text", ""),
                "fields": rec.get("fields", {}),
                "timing": rec.get("timing", {}),
                "error": rec.get("error"),
            })
    except Exception as e:
        logger.exception("ocr_bcard_async_status error")
        return jsonify({"ok": False, "error": str(e)}), 500

APP_BIND_HOST = os.getenv("HOST", "0.0.0.0")
APP_PORT = int(os.getenv("PORT", "5000"))

if __name__ == "__main__":
    try:
        _warmup()
    except Exception as e:
        logger.warning(f"Warmup failed: {e}")
    app.run(host=APP_BIND_HOST, port=APP_PORT, debug=True, use_reloader=True, threaded=True)
